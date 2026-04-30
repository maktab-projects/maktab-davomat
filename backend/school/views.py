from collections import defaultdict
from datetime import datetime, timedelta

from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from rest_framework import generics, status, viewsets
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.models import User
from .models import (
    Classroom,
    SchoolSettings,
    Student,
    StudentAttendance,
    TeacherAttendance,
    get_today_teacher_attendance,
    validate_location_or_raise,
)
from .serializers import (
    BulkAttendanceSerializer,
    CheckInOutSerializer,
    ClassroomSerializer,
    SchoolSettingsSerializer,
    StudentAttendanceSerializer,
    StudentSerializer,
    TeacherAttendanceSerializer,
    UserSerializer,
)


def recent_school_days(limit=14):
    days = []
    cursor = timezone.localdate()
    while len(days) < limit:
        if cursor.weekday() != 6:  # Sunday is ignored
            days.append(cursor)
        cursor -= timedelta(days=1)
    return list(reversed(days))


UZ_WEEKDAYS = {
    0: 'Dushanba',
    1: 'Seshanba',
    2: 'Chorshanba',
    3: 'Payshanba',
    4: 'Juma',
    5: 'Shanba',
    6: 'Yakshanba',
}


def uz_weekday(day):
    return UZ_WEEKDAYS.get(day.weekday(), '')


def format_day_item(day):
    return {
        'date': str(day),
        'label': day.strftime('%d.%m'),
        'weekday': uz_weekday(day),
        'label_full': f"{day.strftime('%d.%m')} {uz_weekday(day)}",
    }


def format_time(value):
    if not value:
        return None
    return timezone.localtime(value).strftime('%H:%M')


def format_datetime(value):
    if not value:
        return None
    return timezone.localtime(value).strftime('%d.%m.%Y %H:%M')


def parse_report_date(value, default):
    parsed = parse_date(value) if value else None
    return parsed or default


def date_range_days(start_date, end_date):
    days = []
    cursor = start_date
    while cursor <= end_date:
        days.append(cursor)
        cursor += timedelta(days=1)
    return days


def format_minutes_to_label(minutes):
    hours = minutes // 60
    mins = minutes % 60
    return f"{hours} soat {mins} daqiqa"

def build_student_status_item(record, student, classroom, selected_date=None):
    note_text = record.note if record else ''
    status_key = record.status if record else 'not_marked'
    status_label_map = {
        StudentAttendance.STATUS_PRESENT: 'Keldi',
        StudentAttendance.STATUS_ABSENT: 'Kelmadi',
        StudentAttendance.STATUS_LATE: 'Kech keldi',
        StudentAttendance.STATUS_EXCUSED: 'Sababli',
        'not_marked': 'Davomat qilinmadi',
    }
    if record and record.status == StudentAttendance.STATUS_EXCUSED:
        reason = note_text or 'Sabab kiritilmagan'
    elif record and record.status == StudentAttendance.STATUS_ABSENT:
        reason = note_text or 'Sabab ko‘rsatilmagan'
    elif record and record.status == StudentAttendance.STATUS_LATE:
        reason = note_text or 'Kech kelgan'
    elif record and record.status == StudentAttendance.STATUS_PRESENT:
        reason = note_text or 'Keldi'
    else:
        reason = 'Davomat belgilanmagan'

    return {
        'id': student.id,
        'student_id': student.id,
        'full_name': student.full_name,
        'status_key': status_key,
        'status': status_label_map.get(status_key, record.get_status_display() if record else 'Davomat qilinmadi'),
        'date': str(record.date) if record else (str(selected_date) if selected_date else None),
        'note': note_text,
        'reason': reason,
        'gender': student.gender,
        'gender_display': student.get_gender_display(),
        'parent_name': student.parent_name,
        'phone_primary': student.phone_primary,
        'phone_secondary': student.phone_secondary,
        'father_full_name': student.father_full_name,
        'father_phone': student.father_phone,
        'mother_full_name': student.mother_full_name,
        'mother_phone': student.mother_phone,
        'address': student.address,
        'classroom_id': classroom.id,
        'classroom_name': classroom.name,
        'teacher_name': classroom.teacher.get_full_name() or classroom.teacher.username,
    }


def school_days_from(start_date, end_date):
    days = []
    cursor = start_date
    while cursor <= end_date:
        if cursor.weekday() != 6:
            days.append(cursor)
        cursor += timedelta(days=1)
    return days


def build_teacher_period_stats(records, school_days):
    records_by_teacher = defaultdict(list)
    for record in records:
        records_by_teacher[record.teacher_id].append(record)

    expected_days = len(school_days)
    stats = {}
    for teacher_id, items in records_by_teacher.items():
        present_days = sum(1 for item in items if item.check_in_time)
        worked_minutes = sum(item.worked_minutes for item in items)
        late_days = sum(1 for item in items if item.is_late)
        stats[teacher_id] = {
            'present_days': present_days,
            'worked_minutes': worked_minutes,
            'worked_label': format_minutes_to_label(worked_minutes),
            'late_days': late_days,
            'expected_days': expected_days,
            'missed_days': max(expected_days - present_days, 0),
        }

    return stats


class SchoolSettingsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        serializer = SchoolSettingsSerializer(SchoolSettings.get_solo())
        return Response(serializer.data)

    def put(self, request):
        if request.user.role not in ['director', 'admin']:
            return Response({'detail': 'Ruxsat yo‘q'}, status=403)
        settings_obj = SchoolSettings.get_solo()
        data = request.data.copy()
        # Director/Admin lokatsiya va ish vaqtini frontenddan o'zgartirmaydi.
        for key in ['latitude', 'longitude', 'allowed_radius_meters', 'work_start_time', 'work_end_time']:
            data.pop(key, None)
        serializer = SchoolSettingsSerializer(settings_obj, data=data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


class ClassroomViewSet(viewsets.ModelViewSet):
    serializer_class = ClassroomSerializer
    permission_classes = [IsAuthenticated]
    search_fields = ['name', 'subject', 'teacher__first_name', 'teacher__last_name']
    filterset_fields = ['teacher', 'subject', 'is_active']

    def get_queryset(self):
        qs = Classroom.objects.select_related('teacher').prefetch_related('students').all()
        user = self.request.user
        if user.role == 'teacher':
            return qs.filter(teacher=user)
        return qs

    def perform_create(self, serializer):
        user = self.request.user
        if user.role == 'teacher':
            serializer.save(teacher=user)
        else:
            teacher_id = self.request.data.get('teacher')
            if not teacher_id:
                raise ValidationError({'teacher': 'Teacher tanlanishi kerak'})
            try:
                teacher = User.objects.get(id=teacher_id, role=User.TEACHER, is_active=True)
            except ObjectDoesNotExist as exc:
                raise ValidationError({'teacher': 'Bunday ustoz topilmadi'}) from exc
            serializer.save(teacher=teacher)


class StudentViewSet(viewsets.ModelViewSet):
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated]
    search_fields = ['first_name', 'last_name', 'parent_name', 'phone_primary']
    filterset_fields = ['classroom', 'gender', 'is_active']

    def get_queryset(self):
        qs = Student.objects.select_related('classroom', 'classroom__teacher').all()
        user = self.request.user
        if user.role == User.TEACHER:
            return qs.filter(classroom__teacher=user)
        return qs

    def _ensure_can_change_student(self, classroom):
        user = self.request.user
        if user.role == User.DIRECTOR:
            raise PermissionDenied('Director o‘quvchi yaratishi yoki o‘zgartirishi mumkin emas. Director faqat ko‘radi va Excel yuklaydi.')
        if user.role == User.TEACHER and classroom.teacher != user:
            raise PermissionDenied('Bu sinf sizga tegishli emas')
        if user.role not in [User.ADMIN, User.TEACHER]:
            raise PermissionDenied('O‘quvchi ma’lumotini o‘zgartirishga ruxsat yo‘q')

    def perform_create(self, serializer):
        classroom = serializer.validated_data['classroom']
        self._ensure_can_change_student(classroom)
        serializer.save()

    def perform_update(self, serializer):
        classroom = serializer.validated_data.get('classroom', serializer.instance.classroom)
        self._ensure_can_change_student(classroom)
        serializer.save()

    def perform_destroy(self, instance):
        self._ensure_can_change_student(instance.classroom)
        instance.is_active = False
        instance.save(update_fields=['is_active'])


class AttendanceListView(generics.ListAPIView):
    serializer_class = StudentAttendanceSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['classroom', 'date', 'status', 'teacher']

    def get_queryset(self):
        qs = StudentAttendance.objects.select_related('student', 'classroom', 'teacher').all()
        user = self.request.user
        if user.role == 'teacher':
            return qs.filter(teacher=user)
        return qs


class BulkAttendanceCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if request.user.role == User.DIRECTOR:
            return Response({'detail': 'Director davomat kiritmaydi, faqat nazorat qiladi'}, status=403)
        serializer = BulkAttendanceSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        classroom_id = serializer.validated_data['classroom_id']
        date_value = serializer.validated_data['date']
        records = serializer.validated_data['records']

        classroom = Classroom.objects.get(id=classroom_id)
        if request.user.role == 'teacher' and classroom.teacher != request.user:
            return Response({'detail': 'Bu sinf sizga tegishli emas'}, status=403)

        created = []
        for record in records:
            student = Student.objects.get(id=record['student_id'], classroom=classroom)
            obj, _ = StudentAttendance.objects.update_or_create(
                student=student,
                date=date_value,
                defaults={
                    'classroom': classroom,
                    'teacher': classroom.teacher,
                    'status': record['status'],
                    'note': record.get('note', ''),
                }
            )
            created.append(obj)
        out = StudentAttendanceSerializer(created, many=True)
        return Response(out.data, status=status.HTTP_201_CREATED)


class TeacherCheckInView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if request.user.role != 'teacher':
            return Response({'detail': 'Faqat ustozlar uchun'}, status=403)
        serializer = CheckInOutSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        lat = serializer.validated_data['latitude']
        lon = serializer.validated_data['longitude']
        is_allowed, distance, settings_obj = validate_location_or_raise(lat, lon)
        if not is_allowed:
            return Response({
                'detail': 'Siz maktab hududida emassiz',
                'distance_meters': distance,
                'allowed_radius_meters': settings_obj.allowed_radius_meters,
            }, status=400)

        record = get_today_teacher_attendance(request.user)
        if record.check_in_time:
            return Response({'detail': 'Bugungi check-in allaqachon qilingan'}, status=400)

        now = timezone.localtime()
        record.check_in_time = now
        record.check_in_latitude = lat
        record.check_in_longitude = lon
        work_start = timezone.make_aware(datetime.combine(timezone.localdate(), settings_obj.work_start_time))
        record.is_late = now > (work_start + timedelta(minutes=settings_obj.late_after_minutes))
        record.save()
        data = TeacherAttendanceSerializer(record).data
        data.update({'detail': 'Ishga keldim saqlandi'})
        return Response(data)


class TeacherCheckOutView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if request.user.role != 'teacher':
            return Response({'detail': 'Faqat ustozlar uchun'}, status=403)
        serializer = CheckInOutSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        lat = serializer.validated_data['latitude']
        lon = serializer.validated_data['longitude']
        is_allowed, distance, settings_obj = validate_location_or_raise(lat, lon)
        if not is_allowed:
            return Response({
                'detail': 'Siz maktab hududida emassiz',
                'distance_meters': distance,
                'allowed_radius_meters': settings_obj.allowed_radius_meters,
            }, status=400)

        record = get_today_teacher_attendance(request.user)
        if not record.check_in_time:
            return Response({'detail': 'Avval Ishga keldim ni bosish kerak'}, status=400)
        if record.check_out_time:
            return Response({'detail': 'Bugungi check-out allaqachon qilingan'}, status=400)

        now = timezone.localtime()
        record.check_out_time = now
        record.check_out_latitude = lat
        record.check_out_longitude = lon
        work_end = timezone.make_aware(datetime.combine(timezone.localdate(), settings_obj.work_end_time))
        record.left_early = now < work_end
        record.calculate_worked_minutes()
        record.save()
        data = TeacherAttendanceSerializer(record).data
        data.update({'detail': 'Ishdan ketdim saqlandi'})
        return Response(data)


class TodayTeacherAttendanceView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.role == 'teacher':
            record = get_today_teacher_attendance(request.user)
            return Response(TeacherAttendanceSerializer(record).data)
        qs = TeacherAttendance.objects.select_related('teacher').filter(date=timezone.localdate())
        return Response(TeacherAttendanceSerializer(qs, many=True).data)


def build_teacher_attendance_report_rows(request):
    user = request.user
    today = timezone.localdate()

    selected_date = parse_report_date(request.query_params.get('date'), today)
    start_date = parse_report_date(request.query_params.get('start_date'), selected_date)
    end_date = parse_report_date(request.query_params.get('end_date'), selected_date)
    if end_date < start_date:
        start_date, end_date = end_date, start_date

    if (end_date - start_date).days > 366:
        end_date = start_date + timedelta(days=366)

    teachers_qs = User.objects.filter(role=User.TEACHER, is_active=True).order_by('first_name', 'last_name', 'username')
    if user.role == User.TEACHER:
        teachers_qs = teachers_qs.filter(id=user.id)
    elif user.role not in [User.DIRECTOR, User.ADMIN]:
        raise PermissionDenied('Ruxsat yo‘q')

    teacher_id = request.query_params.get('teacher')
    if teacher_id and user.role in [User.DIRECTOR, User.ADMIN]:
        teachers_qs = teachers_qs.filter(id=teacher_id)

    teachers = list(teachers_qs.prefetch_related('classrooms'))
    days = date_range_days(start_date, end_date)

    records = TeacherAttendance.objects.filter(
        teacher__in=teachers,
        date__gte=start_date,
        date__lte=end_date,
    ).select_related('teacher')
    record_map = {(record.teacher_id, record.date): record for record in records}

    rows = []
    now = timezone.localtime()
    summary = {
        'total_teachers': len(teachers),
        'total_days': len(days),
        'expected_records': len(teachers) * len(days),
        'present_count': 0,
        'checked_out_count': 0,
        'absent_count': 0,
        'late_count': 0,
        'left_early_count': 0,
        'worked_minutes_total': 0,
        'worked_label_total': format_minutes_to_label(0),
    }

    for day in days:
        for teacher in teachers:
            record = record_map.get((teacher.id, day))
            check_in_time = record.check_in_time if record else None
            check_out_time = record.check_out_time if record else None
            has_check_in = bool(check_in_time)
            has_check_out = bool(check_out_time)

            if record and record.worked_minutes:
                worked_minutes = record.worked_minutes
            elif check_in_time and not check_out_time and day == today:
                worked_minutes = max(int((now - timezone.localtime(check_in_time)).total_seconds() // 60), 0)
            else:
                worked_minutes = 0

            if has_check_in and has_check_out:
                status_label = 'Kelgan va ketgan'
            elif has_check_in:
                status_label = 'Ishda / ketim bosilmagan'
            else:
                status_label = 'Kelmagan'

            if has_check_in:
                summary['present_count'] += 1
            else:
                summary['absent_count'] += 1
            if has_check_out:
                summary['checked_out_count'] += 1
            if record and record.is_late:
                summary['late_count'] += 1
            if record and record.left_early:
                summary['left_early_count'] += 1
            summary['worked_minutes_total'] += worked_minutes

            classroom_names = ', '.join([item.name for item in teacher.classrooms.all()])
            rows.append({
                'id': record.id if record else None,
                'teacher': teacher.id,
                'teacher_name': teacher.get_full_name() or teacher.username,
                'teacher_username': teacher.username,
                'subject': teacher.subject,
                'classrooms': classroom_names,
                'date': str(day),
                'weekday': uz_weekday(day),
                'date_label': f"{day.strftime('%d.%m.%Y')} {uz_weekday(day)}",
                'check_in_time': format_time(check_in_time),
                'check_out_time': format_time(check_out_time),
                'check_in_datetime': format_datetime(check_in_time),
                'check_out_datetime': format_datetime(check_out_time),
                'worked_minutes': worked_minutes,
                'worked_label': format_minutes_to_label(worked_minutes),
                'is_late': bool(record and record.is_late),
                'left_early': bool(record and record.left_early),
                'status': status_label,
                'check_in_location': (
                    f"{record.check_in_latitude}, {record.check_in_longitude}"
                    if record and record.check_in_latitude and record.check_in_longitude else ''
                ),
                'check_out_location': (
                    f"{record.check_out_latitude}, {record.check_out_longitude}"
                    if record and record.check_out_latitude and record.check_out_longitude else ''
                ),
            })

    summary['worked_label_total'] = format_minutes_to_label(summary['worked_minutes_total'])
    return {
        'start_date': str(start_date),
        'end_date': str(end_date),
        'selected_date': str(selected_date),
        'days': [format_day_item(day) for day in days],
        'teachers': [UserSerializer(teacher).data for teacher in teachers],
        'summary': summary,
        'rows': rows,
    }


class TeacherAttendanceReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response(build_teacher_attendance_report_rows(request))


class TeacherAttendanceReportExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        data = build_teacher_attendance_report_rows(request)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Ustozlar davomati'
        headers = [
            '№', 'Ustoz ism familyasi', 'Login', 'Fan', 'Sinflari', 'Sana', 'Hafta kuni',
            'Soat nechida kelgan', 'Soat nechida ketgan', 'Qancha ishlagan',
            'Holat', 'Kechikdimi', 'Erta ketdimi'
        ]
        ws.append(headers)
        header_fill = PatternFill('solid', fgColor='EAF2FF')
        border = Border(bottom=Side(style='thin', color='D9E2F3'))
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = header_fill
            cell.border = border

        for index, row in enumerate(data['rows'], start=1):
            ws.append([
                index,
                row['teacher_name'],
                row['teacher_username'],
                row['subject'],
                row['classrooms'],
                row['date'],
                row['weekday'],
                row['check_in_time'] or '',
                row['check_out_time'] or '',
                row['worked_label'],
                row['status'],
                'Ha' if row['is_late'] else 'Yo‘q',
                'Ha' if row['left_early'] else 'Yo‘q',
            ])

        widths = [6, 28, 18, 18, 32, 14, 14, 20, 20, 20, 26, 12, 12]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="ustozlar-davomati.xlsx"'
        wb.save(response)
        return response


class AttendanceDayFilterView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = timezone.localdate()
        available_days = recent_school_days(30)
        available_dates = {str(day): day for day in available_days}

        requested_date = request.query_params.get('date')
        selected_date = available_dates.get(requested_date)
        if not selected_date:
            try:
                selected_date = datetime.strptime(requested_date, '%Y-%m-%d').date() if requested_date else today
            except (TypeError, ValueError):
                selected_date = today

        classroom_qs = Classroom.objects.select_related('teacher').prefetch_related('students').filter(is_active=True).order_by('name')
        if request.user.role == User.TEACHER:
            classroom_qs = classroom_qs.filter(teacher=request.user)

        classrooms = list(classroom_qs)
        requested_classroom_id = request.query_params.get('classroom')
        selected_classroom = None
        if requested_classroom_id:
            try:
                selected_classroom = next(item for item in classrooms if item.id == int(requested_classroom_id))
            except (StopIteration, TypeError, ValueError):
                selected_classroom = None

        classrooms_for_stats = [selected_classroom] if selected_classroom else classrooms
        classrooms_for_stats = [item for item in classrooms_for_stats if item]

        attendance_qs = StudentAttendance.objects.filter(
            classroom__in=classrooms_for_stats,
            date=selected_date,
        ).select_related('student', 'classroom', 'teacher')
        records = list(attendance_qs)
        records_by_classroom = defaultdict(list)
        for record in records:
            records_by_classroom[record.classroom_id].append(record)

        classroom_items = []
        all_student_rows = []
        all_present_students = []
        all_late_students = []
        all_absent_students = []
        all_excused_students = []
        all_not_marked_students = []
        school_summary = {
            'total_classrooms': 0,
            'total_students': 0,
            'entered_count': 0,
            'present_count': 0,
            'late_count': 0,
            'absent_count': 0,
            'excused_count': 0,
            'unexcused_absent_count': 0,
            'total_absent_count': 0,
            'not_marked_count': 0,
        }

        for classroom in classrooms_for_stats:
            students = list(classroom.students.filter(is_active=True).order_by('first_name', 'last_name'))
            record_list = records_by_classroom.get(classroom.id, [])
            record_map = {record.student_id: record for record in record_list}
            present_count = sum(1 for record in record_list if record.status == StudentAttendance.STATUS_PRESENT)
            late_count = sum(1 for record in record_list if record.status == StudentAttendance.STATUS_LATE)
            absent_count = sum(1 for record in record_list if record.status == StudentAttendance.STATUS_ABSENT)
            excused_count = sum(1 for record in record_list if record.status == StudentAttendance.STATUS_EXCUSED)
            total_students = len(students)
            not_marked_count = max(total_students - len(record_list), 0)

            present_rows = []
            late_rows = []
            absent_rows = []
            excused_rows = []
            not_marked_rows = []
            student_rows = []

            for student in students:
                record = record_map.get(student.id)
                item = build_student_status_item(record, student, classroom, selected_date)
                student_rows.append(item)
                all_student_rows.append(item)

                if record and record.status == StudentAttendance.STATUS_PRESENT:
                    present_rows.append(item)
                    all_present_students.append(item)
                elif record and record.status == StudentAttendance.STATUS_LATE:
                    late_rows.append(item)
                    all_late_students.append(item)
                elif record and record.status == StudentAttendance.STATUS_ABSENT:
                    absent_rows.append(item)
                    all_absent_students.append(item)
                elif record and record.status == StudentAttendance.STATUS_EXCUSED:
                    excused_rows.append(item)
                    all_excused_students.append(item)
                    all_absent_students.append(item)
                else:
                    not_marked_rows.append(item)
                    all_not_marked_students.append(item)

            item = {
                'id': classroom.id,
                'name': classroom.name,
                'subject': classroom.subject,
                'room': classroom.room,
                'lesson_time': classroom.lesson_time,
                'shift': classroom.shift,
                'shift_display': classroom.get_shift_display(),
                'teacher': classroom.teacher_id,
                'teacher_name': classroom.teacher.get_full_name() or classroom.teacher.username,
                'student_count': total_students,
                'entered_count': len(record_list),
                'present_count': present_count,
                'late_count': late_count,
                'absent_count': absent_count,
                'excused_count': excused_count,
                'unexcused_absent_count': absent_count,
                'total_absent_count': absent_count + excused_count,
                'not_marked_count': not_marked_count,
                'attendance_entered': bool(record_list),
                'students': student_rows,
                'present_students': present_rows,
                'late_students': late_rows,
                'absent_students': absent_rows,
                'excused_students': excused_rows,
                'not_marked_students': not_marked_rows,
            }
            classroom_items.append(item)

            school_summary['total_classrooms'] += 1
            school_summary['total_students'] += total_students
            school_summary['entered_count'] += len(record_list)
            school_summary['present_count'] += present_count
            school_summary['late_count'] += late_count
            school_summary['absent_count'] += absent_count
            school_summary['excused_count'] += excused_count
            school_summary['unexcused_absent_count'] += absent_count
            school_summary['total_absent_count'] += absent_count + excused_count
            school_summary['not_marked_count'] += not_marked_count

        selected_classroom_payload = None
        if selected_classroom:
            selected_classroom_payload = next((item for item in classroom_items if item['id'] == selected_classroom.id), None)

        def sort_rows(rows):
            return sorted(rows, key=lambda item: (item['classroom_name'], item['full_name']))

        return Response({
            'days': [format_day_item(day) for day in available_days],
            'selected_date': str(selected_date),
            'selected_weekday': uz_weekday(selected_date),
            'selected_classroom_id': selected_classroom.id if selected_classroom else None,
            'selected_classroom': selected_classroom_payload,
            'classrooms': classroom_items,
            'summary': school_summary,
            'school_summary': school_summary,
            'attendance_entered': school_summary['entered_count'] > 0,
            'students': sort_rows(all_student_rows),
            'present_students': sort_rows(all_present_students),
            'late_students': sort_rows(all_late_students),
            'absent_students': sort_rows(all_absent_students),
            'excused_students': sort_rows(all_excused_students),
            'not_marked_students': sort_rows(all_not_marked_students),
        })


class AttendanceDayExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        report_response = AttendanceDayFilterView().get(request)
        data = report_response.data
        absent_only = str(request.query_params.get('absent_only', '')).lower() in ['1', 'true', 'yes', 'ha']
        rows = data.get('absent_students' if absent_only else 'students', [])

        wb = Workbook()
        ws = wb.active
        ws.title = 'Kelmaganlar' if absent_only else 'Kunlik davomat'
        ws.append([
            '№', 'Sana', 'Hafta kuni', 'Sinf', 'Ustoz', "O'quvchi ism familyasi",
            'Holat', 'Izoh/Sabab', "Tug'ilgan sana", 'Jinsi', "O'quvchi nomeri",
            'Qo‘shimcha nomer', 'Otasining ism familyasi', 'Otasining nomeri',
            'Onasining ism familyasi', 'Onasining nomeri', 'Manzil'
        ])
        header_fill = PatternFill('solid', fgColor='EAF2FF')
        border = Border(bottom=Side(style='thin', color='D9E2F3'))
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = header_fill
            cell.border = border

        for idx, row in enumerate(rows, start=1):
            ws.append([
                idx,
                data.get('selected_date'),
                data.get('selected_weekday'),
                row.get('classroom_name'),
                row.get('teacher_name'),
                row.get('full_name'),
                row.get('status'),
                row.get('reason') or row.get('note'),
                row.get('birth_date'),
                row.get('gender_display'),
                row.get('phone_primary'),
                row.get('phone_secondary'),
                row.get('father_full_name'),
                row.get('father_phone'),
                row.get('mother_full_name'),
                row.get('mother_phone'),
                row.get('address'),
            ])

        widths = [6, 14, 14, 22, 26, 32, 18, 28, 16, 12, 18, 18, 28, 18, 28, 18, 36]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        filename = 'bugun-kelmagan-oquvchilar.xlsx' if absent_only else 'kunlik-davomat.xlsx'
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class DashboardOverviewView(APIView):
    permission_classes = [IsAuthenticated]

    def initial(self, request, *args, **kwargs):
        super().initial(request, *args, **kwargs)
        if request.user.role not in ['director', 'admin']:
            raise PermissionDenied("Faqat direktor va admin ko‘ra oladi")

    def get(self, request):
        today = timezone.localdate()
        recent_days = recent_school_days(14)
        recent_day_items = [format_day_item(day) for day in recent_days]
        weekly_days = recent_school_days(7)
        month_start = today.replace(day=1)
        month_days = school_days_from(month_start, today)

        classroom_qs = Classroom.objects.select_related('teacher').prefetch_related('students').order_by('name')
        teacher_qs = User.objects.filter(role=User.TEACHER, is_active=True).prefetch_related('classrooms__students').order_by('first_name', 'last_name')

        classrooms = classroom_qs.count()
        students = Student.objects.filter(is_active=True).count()
        girls = Student.objects.filter(gender=Student.GENDER_FEMALE, is_active=True).count()
        boys = Student.objects.filter(gender=Student.GENDER_MALE, is_active=True).count()
        teachers = teacher_qs.count()

        today_student_attendance = StudentAttendance.objects.filter(date=today).select_related('student', 'classroom', 'teacher')
        today_records_by_classroom = {}
        for record in today_student_attendance:
            today_records_by_classroom.setdefault(record.classroom_id, []).append(record)

        recent_attendance = StudentAttendance.objects.filter(date__in=recent_days).select_related('student', 'classroom', 'teacher')
        attendance_by_classroom_day = {}
        for record in recent_attendance:
            key = (record.classroom_id, record.date)
            attendance_by_classroom_day.setdefault(key, []).append(record)

        present_students = today_student_attendance.filter(status=StudentAttendance.STATUS_PRESENT).count()
        absent_students = today_student_attendance.filter(status=StudentAttendance.STATUS_ABSENT).count()
        late_students = today_student_attendance.filter(status=StudentAttendance.STATUS_LATE).count()
        excused_students = today_student_attendance.filter(status=StudentAttendance.STATUS_EXCUSED).count()

        today_teacher_attendance = TeacherAttendance.objects.filter(date=today).select_related('teacher')
        teacher_attendance_by_id = {item.teacher_id: item for item in today_teacher_attendance}
        present_teachers = sum(1 for item in teacher_attendance_by_id.values() if item.check_in_time)
        absent_teachers = max(teachers - present_teachers, 0)
        late_teachers = today_teacher_attendance.filter(is_late=True).count()
        total_worked_minutes = today_teacher_attendance.aggregate(total=Sum('worked_minutes'))['total'] or 0

        weekly_teacher_records = list(
            TeacherAttendance.objects.filter(date__in=weekly_days).select_related('teacher')
        )
        monthly_teacher_records = list(
            TeacherAttendance.objects.filter(date__gte=month_start, date__lte=today).select_related('teacher')
        )
        weekly_teacher_stats = build_teacher_period_stats(weekly_teacher_records, weekly_days)
        monthly_teacher_stats = build_teacher_period_stats(monthly_teacher_records, month_days)

        stats_cards = {
            'weekly_present_total': sum(item['present_days'] for item in weekly_teacher_stats.values()),
            'weekly_worked_minutes_total': sum(item['worked_minutes'] for item in weekly_teacher_stats.values()),
            'monthly_present_total': sum(item['present_days'] for item in monthly_teacher_stats.values()),
            'monthly_worked_minutes_total': sum(item['worked_minutes'] for item in monthly_teacher_stats.values()),
        }

        today_absent_students = []
        student_classroom_cards = []
        classroom_stats = []

        for classroom in classroom_qs:
            students_in_class = list(classroom.students.filter(is_active=True).order_by('first_name', 'last_name'))
            record_list = today_records_by_classroom.get(classroom.id, [])
            record_map = {item.student_id: item for item in record_list}
            total_students_in_class = len(students_in_class)

            attendance_present = sum(1 for item in record_list if item.status == StudentAttendance.STATUS_PRESENT)
            attendance_absent = sum(1 for item in record_list if item.status == StudentAttendance.STATUS_ABSENT)
            attendance_late = sum(1 for item in record_list if item.status == StudentAttendance.STATUS_LATE)
            attendance_excused = sum(1 for item in record_list if item.status == StudentAttendance.STATUS_EXCUSED)

            absent_student_rows = []
            for student in students_in_class:
                record = record_map.get(student.id)
                if not record or record.status not in [StudentAttendance.STATUS_ABSENT, StudentAttendance.STATUS_EXCUSED]:
                    continue
                item = build_student_status_item(record, student, classroom)
                absent_student_rows.append(item)
                today_absent_students.append(item)

            recent_history = []
            for day in recent_days:
                records = attendance_by_classroom_day.get((classroom.id, day), [])
                entered_count = len(records)
                if total_students_in_class == 0:
                    status_key = 'empty'
                    status_label = "O‘quvchi yo‘q"
                elif entered_count >= total_students_in_class:
                    status_key = 'done'
                    status_label = 'Davomat kiritildi'
                elif entered_count > 0:
                    status_key = 'partial'
                    status_label = 'Qisman kiritildi'
                else:
                    status_key = 'not_done'
                    status_label = 'Davomat qilinmadi'
                recent_history.append({
                    **format_day_item(day),
                    'status_key': status_key,
                    'status_label': status_label,
                    'entered_count': entered_count,
                    'expected_count': total_students_in_class,
                })

            student_classroom_cards.append({
                'id': classroom.id,
                'name': classroom.name,
                'subject': classroom.subject,
                'teacher_name': classroom.teacher.get_full_name() or classroom.teacher.username,
                'total_students': total_students_in_class,
                'absent_total': len(absent_student_rows),
                'absent_count': attendance_absent,
                'excused_count': attendance_excused,
                'late_count': attendance_late,
                'absent_students': absent_student_rows,
            })

            classroom_stats.append({
                'id': classroom.id,
                'name': classroom.name,
                'subject': classroom.subject,
                'room': classroom.room,
                'schedule_info': classroom.schedule_info,
                'teacher_name': classroom.teacher.get_full_name() or classroom.teacher.username,
                'total_students': total_students_in_class,
                'attendance_present': attendance_present,
                'attendance_absent': attendance_absent,
                'attendance_late': attendance_late,
                'attendance_excused': attendance_excused,
                'absent_students': absent_student_rows,
                'recent_history': recent_history,
            })

        student_classroom_cards.sort(key=lambda item: (-item['absent_total'], item['name']))
        today_absent_students.sort(key=lambda item: (item['classroom_name'], item['full_name']))

        teacher_stats = []
        for teacher in teacher_qs:
            class_details = []
            total_teacher_students = 0
            teacher_classroom_ids = []
            for classroom in teacher.classrooms.all().order_by('name'):
                student_count = classroom.students.filter(is_active=True).count()
                total_teacher_students += student_count
                teacher_classroom_ids.append(classroom.id)
                class_details.append({
                    'id': classroom.id,
                    'name': classroom.name,
                    'subject': classroom.subject,
                    'student_count': student_count,
                })

            recent_history = []
            total_classrooms = len(teacher_classroom_ids)
            for day in recent_days:
                if total_classrooms == 0:
                    status_key = 'empty'
                    status_label = 'Sinf biriktirilmagan'
                    done_classrooms = 0
                else:
                    done_classrooms = sum(1 for classroom_id in teacher_classroom_ids if attendance_by_classroom_day.get((classroom_id, day)))
                    if done_classrooms == total_classrooms:
                        status_key = 'done'
                        status_label = 'Davomat kiritildi'
                    elif done_classrooms > 0:
                        status_key = 'partial'
                        status_label = 'Qisman kiritildi'
                    else:
                        status_key = 'not_done'
                        status_label = 'Davomat qilinmadi'
                recent_history.append({
                    **format_day_item(day),
                    'status_key': status_key,
                    'status_label': status_label,
                    'done_classrooms': done_classrooms,
                    'total_classrooms': total_classrooms,
                })

            teacher_today = teacher_attendance_by_id.get(teacher.id)
            teacher_today_present = bool(teacher_today and teacher_today.check_in_time)
            weekly_stats = weekly_teacher_stats.get(teacher.id, {
                'present_days': 0,
                'worked_minutes': 0,
                'worked_label': format_minutes_to_label(0),
                'late_days': 0,
                'expected_days': len(weekly_days),
                'missed_days': len(weekly_days),
            })
            monthly_stats = monthly_teacher_stats.get(teacher.id, {
                'present_days': 0,
                'worked_minutes': 0,
                'worked_label': format_minutes_to_label(0),
                'late_days': 0,
                'expected_days': len(month_days),
                'missed_days': len(month_days),
            })
            teacher_stats.append({
                'id': teacher.id,
                'first_name': teacher.first_name,
                'last_name': teacher.last_name,
                'subject': teacher.subject,
                'classroom_count': len(class_details),
                'student_count': total_teacher_students,
                'today_present': teacher_today_present,
                'check_in_time': format_time(teacher_today.check_in_time) if teacher_today else None,
                'check_out_time': format_time(teacher_today.check_out_time) if teacher_today else None,
                'worked_minutes': teacher_today.worked_minutes if teacher_today else 0,
                'today_status_label': 'Bugun kelgan' if teacher_today_present else 'Bugun kelmagan',
                'class_details': class_details,
                'recent_history': recent_history,
                'weekly_stats': weekly_stats,
                'monthly_stats': monthly_stats,
            })

        return Response({
            'cards': {
                'classrooms': classrooms,
                'students': students,
                'girls': girls,
                'boys': boys,
                'teachers': teachers,
                'present_students': present_students,
                'absent_students': absent_students,
                'late_students': late_students,
                'excused_students': excused_students,
                'present_teachers': present_teachers,
                'absent_teachers': absent_teachers,
                'late_teachers': late_teachers,
                'total_worked_minutes': total_worked_minutes,
            },
            'recent_school_days': recent_day_items,
            'today_absent_students': today_absent_students,
            'student_classroom_cards': student_classroom_cards,
            'classroom_stats': classroom_stats,
            'teacher_stats': teacher_stats,
            'statistics': {
                'week_range': {
                    'start': str(weekly_days[0]) if weekly_days else None,
                    'end': str(weekly_days[-1]) if weekly_days else None,
                    'expected_days': len(weekly_days),
                },
                'month_range': {
                    'start': str(month_days[0]) if month_days else None,
                    'end': str(month_days[-1]) if month_days else None,
                    'expected_days': len(month_days),
                },
                'cards': stats_cards,
            },
        })


def _classroom_info_queryset_for_user(request):
    qs = Classroom.objects.select_related('teacher').prefetch_related('students').filter(is_active=True).order_by('name')
    user = request.user
    if user.role == User.TEACHER:
        qs = qs.filter(teacher=user)

    classroom_id = request.query_params.get('classroom')
    if classroom_id:
        qs = qs.filter(id=classroom_id)

    if user.role in [User.DIRECTOR, User.ADMIN]:
        teacher_id = request.query_params.get('teacher')
        if teacher_id:
            qs = qs.filter(teacher_id=teacher_id)

    search = request.query_params.get('search')
    if search:
        qs = qs.filter(
            Q(name__icontains=search)
            | Q(subject__icontains=search)
            | Q(teacher__first_name__icontains=search)
            | Q(teacher__last_name__icontains=search)
            | Q(teacher__username__icontains=search)
        )

    return qs


def _student_info_queryset_for_user(request):
    qs = Student.objects.select_related('classroom', 'classroom__teacher').filter(is_active=True)
    user = request.user
    if user.role == User.TEACHER:
        qs = qs.filter(classroom__teacher=user)

    classroom_id = request.query_params.get('classroom')
    if classroom_id:
        qs = qs.filter(classroom_id=classroom_id)

    if user.role in [User.DIRECTOR, User.ADMIN]:
        teacher_id = request.query_params.get('teacher')
        if teacher_id:
            qs = qs.filter(classroom__teacher_id=teacher_id)

    search = request.query_params.get('search')
    if search:
        qs = qs.filter(
            Q(first_name__icontains=search)
            | Q(last_name__icontains=search)
            | Q(phone_primary__icontains=search)
            | Q(father_full_name__icontains=search)
            | Q(father_phone__icontains=search)
            | Q(mother_full_name__icontains=search)
            | Q(mother_phone__icontains=search)
            | Q(address__icontains=search)
            | Q(classroom__name__icontains=search)
        )

    return qs.order_by('classroom__name', 'first_name', 'last_name')


class ClassroomInfoView(APIView):
    """Teacher o'z sinflarini, Director/Admin esa barcha sinflarni ko'radi."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        classrooms = []
        for classroom in _classroom_info_queryset_for_user(request):
            students = classroom.students.filter(is_active=True).order_by('first_name', 'last_name')
            classrooms.append({
                'id': classroom.id,
                'name': classroom.name,
                'subject': classroom.subject,
                'room': classroom.room,
                'lesson_time': classroom.lesson_time,
                'shift': classroom.shift,
                'shift_display': classroom.get_shift_display(),
                'schedule_info': classroom.schedule_info,
                'teacher': classroom.teacher_id,
                'teacher_name': classroom.teacher.get_full_name() or classroom.teacher.username,
                'student_count': students.count(),
                'students': StudentSerializer(students, many=True).data,
            })
        return Response({'classrooms': classrooms})


class MalumotlarView(APIView):
    """Director/Admin: barcha o'quvchilar. Teacher: faqat o'z sinfidagi o'quvchilar."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        student_qs = _student_info_queryset_for_user(request)
        classroom_qs = _classroom_info_queryset_for_user(request)
        students = StudentSerializer(student_qs, many=True).data
        classrooms = []
        for classroom in classroom_qs:
            classroom_students = student_qs.filter(classroom_id=classroom.id)
            classrooms.append({
                'id': classroom.id,
                'name': classroom.name,
                'subject': classroom.subject,
                'room': classroom.room,
                'lesson_time': classroom.lesson_time,
                'shift': classroom.shift,
                'shift_display': classroom.get_shift_display(),
                'schedule_info': classroom.schedule_info,
                'teacher': classroom.teacher_id,
                'teacher_name': classroom.teacher.get_full_name() or classroom.teacher.username,
                'student_count': classroom_students.count(),
                'students': StudentSerializer(classroom_students, many=True).data,
            })
        return Response({
            'title': 'Ma’lumotlar',
            'total_students': student_qs.count(),
            'total_classrooms': classroom_qs.count(),
            'students': students,
            'classrooms': classrooms,
        })


class DirectorMalumotlarView(MalumotlarView):
    def get(self, request):
        if request.user.role not in [User.DIRECTOR, User.ADMIN]:
            return Response({'detail': 'Faqat director yoki admin uchun'}, status=403)
        return super().get(request)


def _write_students_excel_response(request, filename='malumotlar-oquvchilar.xlsx'):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'Ma’lumotlar'
    headers = [
        '№', 'ID', 'Sinf', 'Ustoz', 'Xona', 'Dars vaqti', 'Hafta kuni/Smena',
        "O'quvchi ism familyasi", "Tug'ilgan sana", 'Jinsi', "O'quvchi nomeri",
        "Qo'shimcha nomer", 'Ota/Ona umumiy ismi',
        'Otasining ism familyasi', 'Otasining nomeri',
        'Onasining ism familyasi', 'Onasining nomeri', 'Manzil', 'Holati'
    ]
    ws.append(headers)
    header_fill = PatternFill('solid', fgColor='EAF2FF')
    border = Border(bottom=Side(style='thin', color='D9E2F3'))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = header_fill
        cell.border = border

    for row_index, student in enumerate(_student_info_queryset_for_user(request), start=1):
        classroom = student.classroom
        teacher_name = classroom.teacher.get_full_name() or classroom.teacher.username
        ws.append([
            row_index,
            student.id,
            classroom.name,
            teacher_name,
            classroom.room,
            classroom.lesson_time,
            classroom.get_shift_display(),
            student.full_name,
            student.birth_date.strftime('%d.%m.%Y') if student.birth_date else '',
            student.get_gender_display(),
            student.phone_primary,
            student.phone_secondary,
            student.parent_name,
            student.father_full_name,
            student.father_phone,
            student.mother_full_name,
            student.mother_phone,
            student.address,
            'Faol' if student.is_active else 'Faol emas',
        ])

    widths = [6, 10, 18, 24, 12, 16, 18, 28, 16, 12, 18, 18, 24, 28, 18, 28, 18, 36, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


class ClassroomInfoExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return _write_students_excel_response(request, 'classroom-students.xlsx')


class MalumotlarExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return _write_students_excel_response(request, 'malumotlar-oquvchilar.xlsx')


class DirectorMalumotlarExportView(MalumotlarExportView):
    def get(self, request):
        if request.user.role not in [User.DIRECTOR, User.ADMIN]:
            return Response({'detail': 'Faqat director yoki admin uchun'}, status=403)
        return super().get(request)
